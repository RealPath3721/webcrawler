import requests
from bs4 import BeautifulSoup
from collections import deque
import re

import openpyxl
from openpyxl import Workbook

from datetime import datetime
from tkinter.filedialog import askopenfilename

inputpath = askopenfilename(initialdir='./')
inputbook = openpyxl.load_workbook(inputpath)
inputsheet = inputbook.active
input_max_row = inputsheet.max_row
max_depth = 1
# input_max_row = 11

starting_urls = []

def remove_illegal_characters(text):
    # Define a regular expression pattern to match illegal characters
    illegal_pattern = re.compile(r"[^\x09\x0a\x0d\x20-\x7e]")

    # Remove or replace illegal characters
    cleaned_text = illegal_pattern.sub('', text)
    return cleaned_text



def generate_output_path(inputpath):
    path_array = inputpath.split('/')
    path = ""
    for i in range(0, len(path_array)-1):
        path += path_array[i] + "/"

    curTime = datetime.now()
    timestamp = curTime.strftime("%Y%m%d%H%M%S")
    # path += path_array[-1].split('.')[0] + "_" + timestamp + ".xlsx"
    path += 'output' + "_" + timestamp + ".xlsx"

    return path

outpath = generate_output_path(inputpath)

# Create Starting Urls from .xlsx
for i in range(2, input_max_row + 1):
    cell_obj = inputsheet.cell(row = i, column = 2)
    starting_urls.append(cell_obj.value)
    
outbook = openpyxl.Workbook()
outsheet = outbook.active
outsheet.cell(row=1, column=1).value = "Company"
outsheet.cell(row=1, column=2).value = "Strength"
outsheet.cell(row=1, column=3).value = "Website"
outsheet.cell(row=1, column=4).value = "Triggered Page"
outsheet.cell(row=1, column=5).value = "Links"
outsheet.cell(row=1, column=6).value = "Keyword"
outsheet.cell(row=1, column=7).value = "Strong Positive"
outsheet.cell(row=1, column=8).value = "Positive"


for i in range(2, input_max_row + 1):
    outsheet.cell(row = i, column = 1).value = inputsheet.cell(row = i, column = 1).value
    outsheet.cell(row = i, column = 3).value = inputsheet.cell(row = i, column = 2).value


# Function to check if a URL is valid based on a regular expression pattern
def is_valid_url(url, pattern):
    return re.match(pattern, url)

def is_word_alone(word, txt):
  # Check if the word is alone in the text
  return f" {word} " in f" {txt} "

def is_positive(txt, a_links, iframe_links, url, obj):

    temp_info = obj

    # strong_key = ['Microsoft PowerBI', 'Microsoft Power BI', 
    #               'PowerBI', 'Power BI']
    
    strong_key = ['Microsoft PowerBI', 'Microsoft Power BI', 
                  'PowerBI', 'Power BI', 'Microsoft Power Platform', 
                  'Microsoft technology', 'Microsoft Stack', 
                  'Microsoft 365 Apps']
    medium_key = ['Data Visualization', 'Data Analytics Reporting', 
                  'Business Intelligence', 'Data Analytics', 
                  'Reporting', '%Analytics', 'Market Intelligence', 
                  'Intelligence', '%Analysis', 'Templates']
    link_key = ['https://app.powerbi.com/view', 'https://app.powerbi.com/reportEmbed', 
                'https://powerbicdn.azur']
    strong_positive_texts = ['Pricing', 'Demo', 'Request Demo']
    positive_texts = ['Schedule Demo', 'Reporting & Analytics', 
                      'Analytics', 'Product', 'Products', 'Get']
    for item in strong_positive_texts:
        if is_word_alone(item.upper(), txt.upper()):
            # print(f"Found '{item}' at URL: {url}")
            
            temp_info['spt'].add(item)
            temp_info['spt_counter'] += 1

    for item in positive_texts:
        if is_word_alone(item.upper(), txt.upper()):
            # print(f"Found '{item}' at URL: {url}")

            temp_info['pt'].add(item)
            temp_info['pt_counter'] += 1
    
    for item in strong_key:
        if is_word_alone(item.upper(), txt.upper()):

            print(f"Found '{item}' at URL: {url}")

            temp_info['sk'].add(item)
            temp_info['sk_counter'] += 1

    for item in medium_key:
        if is_word_alone(item.upper(), txt.upper()):

            # print(f"Found '{item}' at URL: {url}")

            temp_info['mk'].add(item)
            temp_info['mk_counter'] += 1
    

    for link in a_links:
        for item in link_key:
            if link != None:
                if item in link:
                    # print(f"Found '{item}' at URL: {url}")

                    temp_info['lk'].add(item)
                    temp_info['lk_counter'] += 1
    
    for link in iframe_links:
        for item in link_key:
            if link != None:
                if item in link:
                    # print(f"Found '{item}' at URL: {url}")
                    temp_info['lk'].add(item)
                    temp_info['lk_counter'] += 1

    temp_info['url'].add(url)

    # print('tmp_info', temp_info)

    return temp_info


def is_negative(txt, url):
    negative = True
    negative_texts = ['Schedule Free Consultation', 'Free Consultation', 
                      'Consulting/Consultants', 'Power BI Services', 
                      'Expertise BI']
    # 'Training', 'Consultation'
    for item in negative_texts:
        if is_word_alone(item.upper(), txt.upper()):
            # print(f"Found '{item}' at URL: {url}")
            negative = False
            break
    return negative

    
def crawl(starting_url, max_depth):
    obj = {
        'spt' : set(),
        'spt_counter' : 0,
        'pt' : set(),
        'pt_counter' : 0,
        'sk' : set(),
        'sk_counter' : 0,
        'mk' : set(),
        'mk_counter' : 0,
        'lk' : set(),
        'lk_counter' : 0,
        'url' : set()
    }
    
    queue = deque([(starting_url, 0)])
    visited = set()
    info = {}

    while queue:
        url, depth = queue.popleft()
        # print(url, depth)
        if url not in visited and depth <= max_depth:
            try:
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'
                }
                # Send a GET request to the webpage
                timeout = 5
                response = requests.get(url, headers = headers, timeout=timeout)
                
                # Parse the HTML content
                soup = BeautifulSoup(response.content, 'html.parser')
                text = soup.get_text()
                # print(text)

                # print(soup.find_all('a'))

                # Find all <a> tags and extract their href attributes
                a_links = [link.get('href') for link in soup.find_all('a')]
                # print(a_links)

                # Find all <iframe> tags and extract their src attributes
                iframe_links = [link.get('src') for link in soup.find_all('iframe')]


                if is_negative(text, url):
                    # print('here1')
                    info = is_positive(text, a_links, iframe_links, url, obj)
                    obj = info

                    for link in a_links:
                        if link != None:
                            if link[0: 4] != 'http':
                                xLink = link[1:]
                                a_links[a_links.index(link)] = f'{url}{xLink}'
                            

                    for link in a_links:
                        if link not in visited:
                            # Enqueue each valid link with an incremented depth
                            queue.append((link, depth + 1))
                        # if is_valid_url(link, url_pattern):
                        #     print(link['href'])
                        #     if link['href'] not in visited:
                        #         # Enqueue each valid link with an incremented depth
                        #         queue.append((link['href'], depth + 1))
                    visited.add(url)
                else:
                    a = 1
                    # queue = False
                    # print('It is negative.')

            except Exception as e:
                err = f"Error crawling {url}: {str(e)}"
                # Handle any exceptions or errors encountered during crawling
                # print(f"Error crawling {url}: {str(e)}")

    return info

tmp = 1
for starting_url in starting_urls:
    # domain = starting_url.split('/')[2]
    # url_pattern = fr'^https?://.*{domain}.*$'
    lk_total = 0
    sk_total = 0
    spt_total = 0
    mk_total = 0
    pt_total = 0
    Triggered_page = ""
    Links = ""
    Keywords = ""
    Strong_positive = ""
    Positive = ""
    status = "Not need"
    info = crawl(starting_url, max_depth)

    if info:
        # if (info['lk_counter'] >= 1) or (info['sk_counter'] >= 1 and info['spt_counter'] >= 1) or ((info['sk_counter'] + info['spt_counter']) >= 2):
        if (info['lk_counter'] >= 1) or (info['sk_counter'] >= 1 and info['spt_counter'] >= 1):
            status = 'Hot'
        elif (info['sk_counter'] >= 1):
            status = 'Medium'
        # elif (info['mk_counter'] >= 2 and info['pt_counter'] >= 1) or ((info['mk_counter'] + info['pt_counter']) >= 3):
        elif (info['mk_counter'] >= 2 and info['pt_counter'] >= 1):

            status = "Potential"
        Triggered_page = ', '.join(str(element) for element in info['url'])
        Links = ', '.join(str(element) for element in info['lk'])
        Keywords = ', '.join(str(element) for element in info['sk']) + ',' + ', '.join(str(element) for element in info['mk'])
        Strong_positive = ', '.join(str(element) for element in info['spt'])
        Positive = ', '.join(str(element) for element in info['pt'])
    # print(starting_url, status)
    outsheet.cell(row = tmp + 1, column = 2).value = status
    outsheet.cell(row = tmp + 1, column = 4).value = remove_illegal_characters(Triggered_page)
    # outsheet.cell(row = tmp + 1, column = 4).value = Triggered_page
    outsheet.cell(row = tmp + 1, column = 5).value = Links
    outsheet.cell(row = tmp + 1, column = 6).value = Keywords
    outsheet.cell(row = tmp + 1, column = 7).value = Strong_positive
    outsheet.cell(row = tmp + 1, column = 8).value = Positive
    
    tmp += 1

outbook.save(outpath)
